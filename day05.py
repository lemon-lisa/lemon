
'''
接口测试
1.用例----接口文档
2,数据获取
3.执行----
4.执行结果和预期结果做比对----断言
5.最终结果写进excel---finresult(passed,failed)
'''
import openpyxl

from lemon_day01.day04 import abc,efg

'''
1,实体功能写出来
2，不固定的东西参数化
3,返回值
'''
import requests
def apitest(url,param):
    header = {'X-Lemonban-Media-Type': 'lemonban.v2',
              'Content-Type': 'application/json'}
    res=requests.post(url=url,json=param,headers=header)
    return res.json()
#
# url='http://8.129.91.152:8766/futureloan/member/register'
# param={
#   "mobile_phone": "15815577988",
#   "pwd": "lemon1234",
#   "type":"0",
#   "reg_name":"管理员用户lemon"
# }

#
# res=apitest(url,param,header)
# print(res)

'''第三方库 openpyxl  可以实现数据的读取和写入'''
''''''
# TODO：读取
# 找到文件  加载进来---loadworkbook---参数是文件名
wb=openpyxl.load_workbook('test_case_api.xlsx')
# print(wb)

# 在哪个表单
sheet=wb['register']
# print(sheet)
#
# #找单元格
cell=sheet.cell(row=17,column=5)
print(cell.value)
#
# # TODO：写入
wb=openpyxl.load_workbook('test_case_api.xlsx')
# print(wb)
#
# # 在哪个表单
sheet=wb['register']
# print(sheet)
#
# #找单元格---一定不能忘了保存----关闭文件
# cell=sheet.cell(row=2,column=8)
# cell.value='胖胖'
# wb.save('test_case_api.xlsx')

# 读的数据   url  param  header  expected
def read_data(filepath,sheet):
    wb=openpyxl.load_workbook(filepath)
    sheet=wb[sheet]
    max_row=sheet.max_row
    # print(max_row)
    # 嵌套字典列表【{}，{}】
    list1 = []
    for i in range(2,max_row+1):
        case=dict(
        id =sheet.cell(row=i, column=1).value,
        url = sheet.cell(row=i, column=5).value,
        parm = sheet.cell(row=i, column=6).value,
        expected = sheet.cell(row=i, column=7).value)
        list1.append(case)
    return list1

def write(filepath,sheet,finresult,row,column):
    print(666)
    #如果你做一个项目列可以固定，行是变化
    wb=openpyxl.load_workbook(filepath)
    sheet=wb[sheet]
    sheet.cell(row=row,column=column).value=finresult
    wb.save(filepath)

# ？？？行是和id有关系的
#  读数据的函数----嵌套字典的列表的测试用例[{用例1需要的数据}，{用例2需要的数据}]
#  参数{用例参数}执行的函数-----执行的结果----
#  写的函数-------依赖于执行结果和用例里面的期望值做比对




